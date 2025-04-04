# -*- coding: utf-8 -*-
import os
import sqlite3
from tkinter import Tk
from tkinter.filedialog import askopenfilename

from loguru import logger
from openpyxl import load_workbook

table_name = "parsing"  # Имя таблицы в базе данных
file_database = "data/data.db"  # Имя файла базы данных


async def opening_the_database():
    """Открытие базы данных"""
    conn = sqlite3.connect("data/data.db")  # Создаем соединение с базой данных
    cursor = conn.cursor()
    return conn, cursor


async def opening_a_files():
    """Открытие файла Excel выбором файла"""
    root = Tk()
    root.withdraw()
    filename = askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    return filename


async def compare_and_rewrite_professions():
    """Изменение от 24.01.2024 Сравнение и перезапись значений профессии в файле Excel счет начинается с 0"""
    conn, cursor = opening_the_database()
    # Открываем выбор файла Excel для чтения данных
    filename = await opening_a_files()
    # Загружаем выбранный файл Excel
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    # Считываем значения из базы данных
    cursor.execute(f"SELECT * FROM {table_name}")
    db_data = cursor.fetchall()
    # Сравниваем значения колонки табельного номера с базой данных и перезаписываем значение профессии в колонку C
    for row in sheet.iter_rows(min_row=5, max_row=1077):
        value_D = str(row[5].value)  # Значение в колонке с которой сравниваются данные
        logger.info(value_D)
        matching_rows = [db_row for db_row in db_data if db_row[0] == value_D]
        if matching_rows:
            profession = matching_rows[0][1]
            row[9].value = profession  # Записываем данные если найдены сходства

    workbook.save(filename)  # Сохраняем изменения в файле Excel
    workbook.close()
    conn.close()  # Закрываем соединение с базой данных


async def parsing_document_1(min_row, max_row, column, column_1) -> None:
    """
    Осуществляет парсинг данных из файла Excel и вставляет их в базу данных SQLite.

    Аргументы:
    :param min_row: Строка, с которой начинается считывание данных.
    :param max_row: Строка, с которой заканчивается считывание данных.
    :param column: Столбец, с которого начинается считывание данных.
    :param column_1: Столбец, с которого начинается считывание данных.
    """
    filename = await opening_a_files()  # Открываем выбор файла Excel для чтения данных
    workbook = load_workbook(filename=filename)  # Загружаем выбранный файл Excel
    sheet = workbook.active
    try:
        os.remove(file_database)  # Удаляем файл базы данных
    except FileNotFoundError:
        logger.info("Файл базы данных не найден")

    conn = sqlite3.connect(file_database)  # Создаем соединение с базой данных
    cursor = conn.cursor()
    # Создаем таблицу в базе данных, если она еще не существует
    cursor.execute(f"CREATE TABLE IF NOT EXISTS {table_name} (table_column_1, table_column_2)")
    # Считываем данные из колонки A и вставляем их в базу данных
    for row in sheet.iter_rows(
            min_row=int(min_row), max_row=int(max_row), values_only=True
    ):
        table_column_1 = str(row[int(column)])  # Преобразуем значение в строку
        table_column_2 = str(row[int(column_1)])  # Преобразуем значение в строку
        # Проверяем, существует ли запись с таким табельным номером в базе данных
        cursor.execute(
            f"SELECT * FROM {table_name} WHERE table_column_1 = ? AND table_column_2 = ?",
            (table_column_1, table_column_2),
        )
        existing_row = cursor.fetchone()
        # Если запись с таким табельным номером не существует, вставляем данные в базу данных
        if existing_row is None:
            cursor.execute(f"INSERT INTO {table_name} VALUES (?, ?)", (table_column_1, table_column_2), )
    # Удаляем повторы по табельному номеру
    cursor.execute(
        f"DELETE FROM {table_name} WHERE rowid NOT IN (SELECT min(rowid) FROM {table_name} GROUP BY table_column_1, table_column_2)")
    # Сохраняем изменения в базе данных и закрываем соединение
    conn.commit()
    conn.close()


if __name__ == "__main__":
    compare_and_rewrite_professions()  # Запускаем функцию для парсинга данных из файла Excel
