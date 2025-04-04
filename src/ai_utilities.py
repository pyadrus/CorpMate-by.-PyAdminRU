import configparser

import openpyxl as op
from gigachat import GigaChat
from loguru import logger

config = configparser.ConfigParser()
config.read("config.ini")
token = config["token"]["token"]


async def gigachat(employee_name):
    # Используйте токен, полученный в личном кабинете из поля Авторизационные данные
    with GigaChat(credentials=token, verify_ssl_certs=False) as giga:
        response = giga.chat(f"Напиши в родительном падеже: {employee_name}")
        print(response.choices[0].message.content)
        return response.choices[0].message.content


async def open_list_gup_docx():
    """Открываем документ с шаблоном"""
    wb = op.load_workbook("../data/list_gup/Списочный_состав.xlsx")  # открываем файл
    sheet = wb.active  # открываем активную таблицу
    current_row = 6  # Начальная строка   1077
    for row in sheet.iter_rows(min_row=6, max_row=1077, values_only=True):
        number = str(row[30])  # Считываем значение в колонке
        logger.info(number)
        sheet.cell(
            row=current_row, column=35, value=gigachat(number)
        )  # Записываем значение в ячейку
        current_row += 1  # Переходим к следующей строке
    wb.save("list_gup/Списочный_состав.xlsx")


if __name__ == "__main__":
    open_list_gup_docx()
